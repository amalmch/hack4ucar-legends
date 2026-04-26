"""
reports/alert_generator.py
───────────────────────────
Layer 4 — Automated Reports & Alerts via Mistral 7B

Two output types:
  1. alert_message(...)   → Short alert (1–3 sentences) for dashboard/email
  2. generate_report(...) → Full PDF/Excel structured report

LLM backend: Ollama (local) or Hugging Face Inference API
"""
from __future__ import annotations

import json
import os
from datetime import datetime
from typing import Literal, Optional

import requests
from jinja2 import Environment, FileSystemLoader, select_autoescape
from loguru import logger

from config import get_config

cfg = get_config()


# ══════════════════════════════════════════════════════════
#  Mistral LLM Client
# ══════════════════════════════════════════════════════════

class MistralClient:
    """
    Thin wrapper around Mistral 7B via:
      - Ollama (local): POST http://localhost:11434/api/generate
      - Hugging Face Inference API: requires HF_API_KEY
    """

    def __init__(self):
        self.backend = cfg.MISTRAL_BACKEND
        self.model = cfg.MISTRAL_MODEL

    def generate(self, prompt: str, max_tokens: int = 512, temperature: float = 0.3) -> str:
        if self.backend == "ollama":
            return self._ollama(prompt, max_tokens, temperature)
        elif self.backend == "hf":
            return self._huggingface(prompt, max_tokens, temperature)
        else:
            raise ValueError(f"Unknown MISTRAL_BACKEND: {self.backend}")

    def _ollama(self, prompt: str, max_tokens: int, temperature: float) -> str:
        url = f"{cfg.OLLAMA_URL}/api/generate"
        payload = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
            "options": {"num_predict": max_tokens, "temperature": temperature},
        }
        try:
            resp = requests.post(url, json=payload, timeout=60)
            resp.raise_for_status()
            return resp.json().get("response", "").strip()
        except requests.RequestException as e:
            logger.error(f"Ollama request failed: {e}")
            return "[Erreur LLM — modèle indisponible]"

    def _huggingface(self, prompt: str, max_tokens: int, temperature: float) -> str:
        api_url = f"https://api-inference.huggingface.co/models/mistralai/Mistral-7B-Instruct-v0.2"
        headers = {"Authorization": f"Bearer {cfg.HF_API_KEY}"}
        payload = {
            "inputs": prompt,
            "parameters": {
                "max_new_tokens": max_tokens,
                "temperature": temperature,
                "return_full_text": False,
            },
        }
        try:
            resp = requests.post(api_url, headers=headers, json=payload, timeout=90)
            resp.raise_for_status()
            data = resp.json()
            if isinstance(data, list) and data:
                return data[0].get("generated_text", "").strip()
            return str(data)
        except requests.RequestException as e:
            logger.error(f"HuggingFace request failed: {e}")
            return "[Erreur LLM — modèle indisponible]"


# ══════════════════════════════════════════════════════════
#  Alert Generator
# ══════════════════════════════════════════════════════════

_llm = MistralClient()


SEVERITY_CONTEXT = {
    "critical": "CRITIQUE — action immédiate requise",
    "warning":  "AVERTISSEMENT — surveillance renforcée recommandée",
    "normal":   "NORMAL — aucune action requise",
}

METRIC_CONTEXT_FR = {
    "taux_reussite":      "taux de réussite des étudiants",
    "taux_abandon":       "taux d'abandon des étudiants",
    "budget_execute_pct": "taux d'exécution budgétaire",
    "nb_inscriptions":    "nombre d'inscriptions",
    "gender_ratio":       "ratio de genre",
}


def alert_message(
    institution_name: str,
    metric: str,
    current_value: float,
    severity: Literal["critical", "warning", "normal"],
    xai_summary: Optional[str] = None,
    use_llm: bool = True,
) -> dict:
    """
    Generate a short alert message in French.

    Returns:
        {
          "institution": str,
          "metric": str,
          "severity": str,
          "message": str,        # LLM-generated or template
          "timestamp": str,
          "action_required": bool
        }
    """
    metric_label = METRIC_CONTEXT_FR.get(metric, metric)
    severity_label = SEVERITY_CONTEXT.get(severity, severity)
    value_str = f"{current_value:.1%}" if current_value <= 1.5 else f"{current_value:.1f}"

    if use_llm:
        prompt = _build_alert_prompt(
            institution_name, metric_label, value_str, severity_label, xai_summary
        )
        message = _llm.generate(prompt, max_tokens=150, temperature=0.2)
    else:
        # Template fallback (no LLM)
        message = (
            f"⚠️ {severity_label} détecté à {institution_name} : "
            f"le {metric_label} est de {value_str}."
        )
        if xai_summary:
            message += f" {xai_summary}"

    return {
        "institution": institution_name,
        "metric": metric,
        "metric_label_fr": metric_label,
        "severity": severity,
        "value": current_value,
        "message": message,
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "action_required": severity in ("critical", "warning"),
    }


def _build_alert_prompt(
    institution: str, metric_label: str, value: str, severity: str, xai: Optional[str]
) -> str:
    xai_part = f"\nCause principale identifiée par l'IA : {xai}" if xai else ""
    return f"""Tu es un assistant d'analyse éducative pour l'UCAR (Union des Universités de Tunisie).
Rédige une alerte courte (2-3 phrases maximum) en français pour un administrateur UCAR.
Sois précis, professionnel et orienté action.

Établissement : {institution}
Indicateur : {metric_label}
Valeur actuelle : {value}
Niveau de gravité : {severity}{xai_part}

Alerte :"""


# ══════════════════════════════════════════════════════════
#  Report Generator
# ══════════════════════════════════════════════════════════

def generate_report(
    kpi_data: dict,
    anomalies: list[dict],
    predictions: list[dict],
    institution_name: str = "UCAR Global",
    format: Literal["pdf", "excel", "json"] = "json",
    use_llm: bool = True,
) -> dict:
    """
    Generate a full structured report in French.

    Args:
        kpi_data:         Latest KPI snapshot (from Binôme A)
        anomalies:        List of detected anomalies with SHAP explanations
        predictions:      List of forecasts per metric
        institution_name: Target institution or "UCAR Global"
        format:           Output format (json = structured dict, pdf/excel = file path)
        use_llm:          Whether to use Mistral 7B for narrative sections

    Returns:
        {
          "title": str,
          "institution": str,
          "generated_at": str,
          "executive_summary": str,
          "kpi_snapshot": dict,
          "anomaly_section": dict,
          "prediction_section": dict,
          "recommendations": list[str],
          "file_path": str | None      # set if format is pdf or excel
        }
    """
    now = datetime.utcnow()
    report = {
        "title": f"Rapport UCAR — {institution_name}",
        "institution": institution_name,
        "generated_at": now.isoformat() + "Z",
        "period": now.strftime("%B %Y"),
        "kpi_snapshot": kpi_data,
        "anomaly_section": _build_anomaly_section(anomalies),
        "prediction_section": _build_prediction_section(predictions),
        "recommendations": [],
        "executive_summary": "",
        "file_path": None,
    }

    # ── LLM-generated narrative ───────────────────────────
    if use_llm:
        report["executive_summary"] = _generate_executive_summary(
            institution_name, kpi_data, anomalies, predictions
        )
        report["recommendations"] = _generate_recommendations(
            anomalies, predictions
        )
    else:
        report["executive_summary"] = _template_summary(institution_name, anomalies, predictions)
        report["recommendations"] = _template_recommendations(anomalies, predictions)

    # ── File export ───────────────────────────────────────
    if format == "pdf":
        report["file_path"] = _export_pdf(report)
    elif format == "excel":
        report["file_path"] = _export_excel(report, kpi_data, anomalies, predictions)

    logger.info(f"Report generated: {institution_name} [{format}]")
    return report


def _build_anomaly_section(anomalies: list[dict]) -> dict:
    critical = [a for a in anomalies if a.get("severity") == "critical"]
    warnings  = [a for a in anomalies if a.get("severity") == "warning"]
    return {
        "total": len(anomalies),
        "critical_count": len(critical),
        "warning_count": len(warnings),
        "details": anomalies[:20],  # Cap for report size
    }


def _build_prediction_section(predictions: list[dict]) -> dict:
    alerts = [p for p in predictions if p.get("alert")]
    return {
        "metrics_forecasted": len(predictions),
        "alerts_forecasted": len(alerts),
        "details": predictions,
    }


def _generate_executive_summary(
    institution: str, kpi: dict, anomalies: list, predictions: list
) -> str:
    n_anomalies = len(anomalies)
    n_critical = sum(1 for a in anomalies if a.get("severity") == "critical")
    n_pred_alerts = sum(1 for p in predictions if p.get("alert"))

    kpi_str = json.dumps({k: round(v, 3) if isinstance(v, float) else v
                          for k, v in (kpi.items() if isinstance(kpi, dict) else {}.items())},
                         ensure_ascii=False)

    prompt = f"""Tu es analyste institutionnel pour l'UCAR. Rédige un résumé exécutif concis (4-6 phrases) en français pour le rapport mensuel.

Établissement : {institution}
KPIs actuels : {kpi_str}
Anomalies détectées : {n_anomalies} dont {n_critical} critiques
Alertes de prédiction : {n_pred_alerts} indicateurs à surveiller

Résumé exécutif :"""

    return _llm.generate(prompt, max_tokens=300, temperature=0.3)


def _generate_recommendations(anomalies: list, predictions: list) -> list[str]:
    critical = [a for a in anomalies if a.get("severity") == "critical"]
    if not critical and not any(p.get("alert") for p in predictions):
        return ["Aucune action corrective urgente requise. Maintenir le suivi régulier."]

    issues = []
    for a in critical[:3]:
        xai = a.get("explanation", {}).get("summary_fr", "")
        issues.append(f"- Anomalie critique : {xai or a.get('metric', 'KPI inconnu')}")
    for p in predictions:
        if p.get("alert"):
            issues.append(f"- Prédiction défavorable : {METRIC_CONTEXT_FR.get(p['metric'], p['metric'])}")

    issues_str = "\n".join(issues)
    prompt = f"""En tant qu'expert en gestion universitaire, propose 3 à 5 recommandations concrètes et actionnables en français, basées sur ces problèmes identifiés :

{issues_str}

Recommandations (liste numérotée) :"""

    text = _llm.generate(prompt, max_tokens=400, temperature=0.4)
    # Parse numbered list
    lines = [l.strip() for l in text.split("\n") if l.strip() and l.strip()[0].isdigit()]
    return lines if lines else [text]


def _template_summary(institution: str, anomalies: list, predictions: list) -> str:
    n = len(anomalies)
    return (
        f"Rapport mensuel de {institution}. "
        f"{n} anomalie(s) détectée(s) sur la période analysée. "
        f"Les prédictions indiquent {'des risques à surveiller' if any(p.get('alert') for p in predictions) else 'une trajectoire stable'}."
    )


def _template_recommendations(anomalies: list, predictions: list) -> list[str]:
    recs = []
    if any(a.get("severity") == "critical" for a in anomalies):
        recs.append("Investiguer immédiatement les anomalies critiques signalées.")
    if any(p.get("alert") for p in predictions):
        recs.append("Mettre en place un plan d'action préventif pour les KPIs en alerte prédictive.")
    recs.append("Effectuer un suivi hebdomadaire des indicateurs signalés.")
    return recs


def _export_pdf(report: dict) -> Optional[str]:
    """Export report to PDF using ReportLab."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        import os

        os.makedirs("./exports", exist_ok=True)
        filename = f"./exports/rapport_{report['institution'].replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"

        doc = SimpleDocTemplate(filename, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                     fontSize=18, spaceAfter=12)
        story.append(Paragraph(report["title"], title_style))
        story.append(Paragraph(f"Généré le : {report['generated_at'][:10]}", styles["Normal"]))
        story.append(Spacer(1, 0.5*cm))

        # Executive summary
        story.append(Paragraph("Résumé Exécutif", styles["Heading2"]))
        story.append(Paragraph(report.get("executive_summary", ""), styles["Normal"]))
        story.append(Spacer(1, 0.5*cm))

        # Anomaly section
        anom = report.get("anomaly_section", {})
        story.append(Paragraph("Anomalies Détectées", styles["Heading2"]))
        story.append(Paragraph(
            f"Total : {anom.get('total', 0)} — Critiques : {anom.get('critical_count', 0)} — Avertissements : {anom.get('warning_count', 0)}",
            styles["Normal"]
        ))
        story.append(Spacer(1, 0.3*cm))

        # Recommendations
        story.append(Paragraph("Recommandations", styles["Heading2"]))
        for rec in report.get("recommendations", []):
            story.append(Paragraph(f"• {rec}", styles["Normal"]))

        doc.build(story)
        logger.info(f"PDF exported: {filename}")
        return filename

    except Exception as e:
        logger.error(f"PDF export failed: {e}")
        return None


def _export_excel(report: dict, kpi_data: dict, anomalies: list, predictions: list) -> Optional[str]:
    """Export report to Excel using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import os

        os.makedirs("./exports", exist_ok=True)
        filename = f"./exports/rapport_{report['institution'].replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"

        wb = openpyxl.Workbook()

        # ── Sheet 1: KPI Snapshot ─────────────────────────
        ws_kpi = wb.active
        ws_kpi.title = "KPIs"
        ws_kpi.append(["Indicateur", "Valeur"])
        if isinstance(kpi_data, dict):
            for k, v in kpi_data.items():
                ws_kpi.append([k, v])

        # ── Sheet 2: Anomalies ────────────────────────────
        ws_anom = wb.create_sheet("Anomalies")
        ws_anom.append(["Institution", "Métrique", "Sévérité", "Explication"])
        for a in anomalies:
            ws_anom.append([
                a.get("institution_id", ""),
                a.get("metric", ""),
                a.get("severity", ""),
                a.get("explanation", {}).get("summary_fr", ""),
            ])

        # ── Sheet 3: Predictions ──────────────────────────
        ws_pred = wb.create_sheet("Prédictions")
        ws_pred.append(["Métrique", "Méthode", "Tendance", "Alerte", "Valeur prédite (fin)"])
        for p in predictions:
            ws_pred.append([
                p.get("metric", ""),
                p.get("method", ""),
                p.get("trend", ""),
                "Oui" if p.get("alert") else "Non",
                p.get("predicted_end", ""),
            ])

        wb.save(filename)
        logger.info(f"Excel exported: {filename}")
        return filename

    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return None
